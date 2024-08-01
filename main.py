import streamlit as st
import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook
# from openpyxl.writer.excel import save_virtual_workbook
import numpy as np
import requests
from PIL import Image
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime

st.title("Formatar arquivos RNCAC")

file_upload = st.file_uploader("Anexe o arquivo original baixado do monday", type='xlsx')

def is_invalid_date(date):
    # Define uma data sentinela que representa uma data indefinida
    sentinel_date = datetime(1900, 1, 1)
    return date == sentinel_date

if file_upload is not None:
    # Salva o arquivo temporário
    temp_file = 'arquivo_formatado.xlsx'
    # temp_file = r"PAINEL_DE_ACOMPANHAMENTO_RNC_1722521575.xlsx"

    with open(temp_file, 'wb') as file:
        file.write(file_upload.getvalue())

    font_bold = Font(bold=True)

    excelOriginal = pd.read_excel(temp_file)
    # temp_file = r"RNC_1706789367.xlsx"
    
    id = excelOriginal['Unnamed: 2'][4] #B7
    # item = excelOriginal['RNCAC'][4] 
    # log_criacao = excelOriginal['Unnamed: 3'][4]
    # origem = excelOriginal['Unnamed: 4'][4]
    setor = excelOriginal['Unnamed: 10'][4] #D7
    # conjunto = excelOriginal['Unnamed: 6'][4] 
    responsavel = excelOriginal['Unnamed: 11'][4] #G8 G15
    # gravidade = excelOriginal['Unnamed: 10'][4]
    # prioridade = excelOriginal['Unnamed: 11'][4]
    status = excelOriginal['Unnamed: 8'][4] 
    # ultima_atualizacao = excelOriginal['Unnamed: 16'][4]
    
    data = excelOriginal['Unnamed: 6'][4]    
    if not isinstance(data, datetime):
        data = ''
    else:
        data = data.strftime('%d/%m/%Y')

    descricao = excelOriginal['PAINEL DE ACOMPANHAMENTO RNC'][4] #A11
    
    try:
        arquivos = excelOriginal['Unnamed: 14'][4]
    except:
        arquivos = ''

    # arquivos_split = arquivos.split()
    # foto = arquivos_split[0].replace(",","")

    # conclusao = excelOriginal['Unnamed: 14'][4]
    acao_contencao = excelOriginal['Unnamed: 15'][4] #A17

    excelOriginal['Unnamed: 19'] = excelOriginal['Unnamed: 19'].fillna('N/A')
    excelOriginal['Unnamed: 20'] = excelOriginal['Unnamed: 20'].fillna('N/A')
    excelOriginal['Unnamed: 21'] = excelOriginal['Unnamed: 21'].fillna('N/A')
    excelOriginal['Unnamed: 22'] = excelOriginal['Unnamed: 22'].fillna('N/A')
    excelOriginal['Unnamed: 23'] = excelOriginal['Unnamed: 23'].fillna('N/A')
    excelOriginal['Unnamed: 24'] = excelOriginal['Unnamed: 24'].fillna('N/A')

    maquina = excelOriginal['Unnamed: 19'][4]
    mao_de_obra = excelOriginal['Unnamed: 20'][4] #B25
    materia_prima = excelOriginal['Unnamed: 21'][4] #B26
    medicao = excelOriginal['Unnamed: 22'][4] #B27
    metodo = excelOriginal['Unnamed: 23'][4] #B28
    meio_ambiente = excelOriginal['Unnamed: 24'][4] #B29
    # encerradoEm = excelOriginal['Unnamed: 25'][4].strftime("%d/%m/%Y")
    responsavelUltimo = excelOriginal['Unnamed: 30'][4]
    participantes = excelOriginal['Unnamed: 26'][4] #C30

    encerradoEm = excelOriginal['Unnamed: 29'][4]
    if not isinstance(encerradoEm, datetime):
        encerradoEm = ''
    else:
        encerradoEm = encerradoEm.strftime('%d/%m/%Y')  

    conclusao = excelOriginal['Unnamed: 28'][4]
    item_norma = excelOriginal['Unnamed: 12'][4]
    avaliacao = excelOriginal['Unnamed: 27'][4]
    status2 = excelOriginal['Unnamed: 17'][4]

    data2 = excelOriginal['Unnamed: 18'][4]
    if not isinstance(data2, datetime):
        data2 = ''
    else:
        data2 = data2.strftime('%d/%m/%Y')  

    data3etapa = excelOriginal['Unnamed: 25'][4]
    if not isinstance(data3etapa, datetime):
        data3etapa = ''
    else:
        data3etapa = data3etapa.strftime('%d/%m/%Y')  

    responsavel2 = excelOriginal['Unnamed: 16'][4]
    conjuntoAtividade = excelOriginal['Unnamed: 13'].fillna('N/A')[4]

    if not np.isnan(avaliacao):
        acao_eficaz = avaliacao * '★'
        acao_eficaz_2 = 'Sim'
    else:
        acao_eficaz = ''
        acao_eficaz_2 = 'Não'

    excelOriginal_cortado = excelOriginal.iloc[5:].reset_index(drop=True)
    # excelOriginal_cortado = excelOriginal_cortado.set_axis(excelOriginal_cortado.iloc[0], axis='columns', inplace=False)
    excelOriginal_cortado = excelOriginal_cortado.set_axis(excelOriginal_cortado.iloc[0], axis='columns')
    excelOriginal_cortado = excelOriginal_cortado[1:].reset_index(drop=True)

    wb = load_workbook('modelo_final_v2.xlsx')
    ws = wb.active
    
    ws['B8'] = id
    ws['D8'] = setor
    ws['G8'] = responsavel
    ws['G15'] = responsavel2
    ws['A12'] = descricao
    ws['A17'] = acao_contencao
    ws['B24'] = maquina
    ws['B25'] = mao_de_obra
    ws['B26'] = materia_prima
    ws['B27'] = medicao
    ws['B28'] = metodo
    ws['B29'] = meio_ambiente
    ws['C30'] = participantes
    ws['G7'] = status
    ws['I8'] = data
    ws['I14'] = data2
    ws['I22'] = data3etapa
    ws['A55'] = conclusao
    ws['E50'] = acao_eficaz
    ws['A53'] = acao_eficaz_2
    ws['G9'] = item_norma
    ws['G14'] = status2
    ws['G11'] = conjuntoAtividade
    ws['G22'] = responsavel2
    ws['G50'] = encerradoEm
    ws['G51'] = responsavelUltimo

    # try:
    #     ws['i22'] = dataPenultima
    # except:
    #     ws['i22'] = ''
    
    # ws['B14'].font = font_bold

    # ws['B15'] = prioridade
    # ws['B15'].font = font_bold

    # ws['B16'] = status
    # ws['B16'].font = font_bold

    # ws['A20'] = descricao
    # ws['A18'] = arquivos
    # ws['D45'] = ultima_atualizacao

    # ws['A42'] = conclusao
    # ws['D44'] = avaliacao

    try:
        excelOriginal_cortado['Previsão'] = pd.to_datetime(excelOriginal_cortado['Previsão'], errors='ignore')
        excelOriginal_cortado['Previsão'] = excelOriginal_cortado['Previsão'].dt.strftime('%d/%m/%Y')
        excelOriginal_cortado['Previsão'] = excelOriginal_cortado['Previsão'].str.replace("-","/")
        excelOriginal_cortado = excelOriginal_cortado.fillna('')
    except:
        excelOriginal['Previsão'] = ''
    
    try:
        excelOriginal_cortado['Conclusão'] = pd.to_datetime(excelOriginal_cortado['Conclusão'], errors='ignore')
        excelOriginal_cortado['Conclusão'] = excelOriginal_cortado['Conclusão'].dt.strftime('%d/%m/%Y')
        excelOriginal_cortado['Conclusão'] = excelOriginal_cortado['Conclusão'].str.replace("-","/")
        excelOriginal_cortado = excelOriginal_cortado.fillna('')
    except:
        excelOriginal['Conclusão'] = ''

    try:
        excelOriginal_cortado['Previsão - End'] = pd.to_datetime(excelOriginal_cortado['Previsão - End'], errors='ignore')
        excelOriginal_cortado['Previsão - End'] = excelOriginal_cortado['Previsão - End'].dt.strftime('%d/%m/%Y')
        excelOriginal_cortado['Previsão - End'] = excelOriginal_cortado['Previsão - End'].str.replace("-","/")
        excelOriginal_cortado = excelOriginal_cortado.fillna('')
    except:
        excelOriginal['Previsão - End'] = ''
        
    u = 35
    ultimaLinha = len(excelOriginal_cortado) + u-1

    df_status = excelOriginal_cortado.iloc[:,5:6]

    # excelOriginal_cortado = excelOriginal_cortado.fillna('N/A')

    excelOriginal_cortado = excelOriginal_cortado.fillna('')

    # try:

        #while u < ultimaLinha:
    for i in range(0,len(excelOriginal_cortado)-1):
        ws['A' + str(u)] = excelOriginal_cortado['Name'][i]
        ws['F' + str(u)] = excelOriginal_cortado['Previsão - End'][i]
        ws['E' + str(u)] = excelOriginal_cortado['Responsável'][i]
        ws['H' + str(u)] = excelOriginal_cortado['Conclusão'][i]
        u += 1
    # except:
    #     pass

    wb.template = False
    wb.save(temp_file)

    # Crie um botão de download
    download_button = st.download_button(
        label="Clique aqui para baixar o arquivo formatado",
        data=open(temp_file, 'rb').read(),
        file_name=temp_file,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Exibe o botão de download
    st.write(download_button)
